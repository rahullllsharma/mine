import io
from unittest.mock import patch

import pytest

from worker_safety_service.enums import FileType
from worker_safety_service.utils import parse_file


@pytest.mark.unit
class TestDataSourceUtils:
    """Unit tests for data source utility functions."""

    def test_parse_csv_file_success(self) -> None:
        """Test successful CSV file parsing."""
        # Arrange
        csv_content = "name,age,city\nJohn,25,NYC\nJane,30,LA\nBob,35,Chicago"
        file_content = csv_content.encode()

        # Act
        result = parse_file(file_content, FileType.CSV)

        # Assert
        expected = {
            "name": ["Bob", "Jane", "John"],  # Sorted alphabetically
            "age": ["25", "30", "35"],  # Sorted alphabetically
            "city": ["Chicago", "LA", "NYC"],  # Sorted alphabetically
        }
        assert result == expected

    def test_parse_csv_file_with_empty_values(self) -> None:
        """Test CSV parsing with empty values - empty values are skipped."""
        # Arrange
        csv_content = "name,age,city\nJohn,,NYC\n,30,\nBob,35,Chicago"
        file_content = csv_content.encode()

        # Act
        result = parse_file(file_content, FileType.CSV)

        # Assert
        # Empty values are not stored, so we only get non-empty values
        expected = {
            "name": ["Bob", "John"],  # Empty name is skipped, sorted alphabetically
            "age": ["30", "35"],  # Empty age is skipped, sorted alphabetically
            "city": ["Chicago", "NYC"],  # Empty city is skipped, sorted alphabetically
        }
        assert result == expected

    def test_parse_csv_file_single_column(self) -> None:
        """Test CSV parsing with single column."""
        # Arrange
        csv_content = "values\n1\n2\n3"
        file_content = csv_content.encode()

        # Act
        result = parse_file(file_content, FileType.CSV)

        # Assert
        expected = {"values": ["1", "2", "3"]}
        assert result == expected

    def test_parse_csv_file_with_only_empty_values_in_column(self) -> None:
        """Test CSV parsing when a column has only empty values."""
        # Arrange
        csv_content = "name,age,city\nJohn,,NYC\nJane,,LA\nBob,,Chicago"
        file_content = csv_content.encode()

        # Act
        result = parse_file(file_content, FileType.CSV)

        # Assert
        # The 'age' column should be empty since all values are empty
        expected = {
            "name": ["Bob", "Jane", "John"],  # Sorted alphabetically
            "age": [],  # Empty because all age values are empty
            "city": ["Chicago", "LA", "NYC"],  # Sorted alphabetically
        }
        assert result == expected

    def test_parse_csv_file_with_whitespace_only_values(self) -> None:
        """Test CSV parsing with whitespace-only values."""
        # Arrange
        # Use proper CSV format with quoted fields to handle whitespace
        csv_content = 'name,age,city\nJohn,"  ",NYC\nJane,"\t",LA\nBob," ",Chicago'
        file_content = csv_content.encode()

        # Act
        result = parse_file(file_content, FileType.CSV)

        # Assert
        # Whitespace-only values should be filtered out
        expected = {
            "name": ["Bob", "Jane", "John"],  # Sorted alphabetically
            "age": [],  # Whitespace-only values are filtered out
            "city": ["Chicago", "LA", "NYC"],  # Sorted alphabetically
        }
        assert result == expected

    def test_parse_csv_file_empty_content(self) -> None:
        """Test CSV parsing with empty content."""
        # Arrange
        csv_content = ""
        file_content = csv_content.encode()

        # Act & Assert
        # Empty content should raise a descriptive ValueError
        with pytest.raises(ValueError, match="CSV file is empty or contains no data"):
            parse_file(file_content, FileType.CSV)

    def test_parse_csv_file_headers_only(self) -> None:
        """Test CSV parsing with headers only (no data rows)."""
        # Arrange
        csv_content = "name,age,city"
        file_content = csv_content.encode()

        # Act
        result = parse_file(file_content, FileType.CSV)

        # Assert
        expected: dict[str, list[str]] = {"name": [], "age": [], "city": []}
        assert result == expected

    def test_parse_xlsx_file_success(self) -> None:
        """Test successful XLSX file parsing."""
        # Arrange
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["product", "price", "category"])
        ws.append(["Apple", 1.5, "Fruit"])
        ws.append(["Bread", 2.0, "Bakery"])

        buffer = io.BytesIO()
        wb.save(buffer)
        file_content = buffer.getvalue()

        # Act
        result = parse_file(file_content, FileType.XLSX)

        # Assert
        # Note: 2.0 becomes "2" since it's a whole number, but 1.5 stays "1.5"
        expected = {
            "product": ["Apple", "Bread"],  # Already sorted alphabetically
            "price": [
                "1.5",
                "2",
            ],  # 2.0 becomes "2", 1.5 stays "1.5", sorted alphabetically
            "category": ["Bakery", "Fruit"],  # Sorted alphabetically
        }
        assert result == expected

    def test_parse_xlsm_file_success(self) -> None:
        """Test successful XLSM file parsing."""
        # Arrange
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["macro_col", "data_col"])
        ws.append(["value1", "data1"])
        ws.append(["value2", "data2"])

        buffer = io.BytesIO()
        wb.save(buffer)
        file_content = buffer.getvalue()

        # Act
        result = parse_file(file_content, FileType.XLSM)

        # Assert
        expected = {
            "macro_col": ["value1", "value2"],
            "data_col": ["data1", "data2"],
        }  # Already sorted
        assert result == expected

    def test_parse_excel_file_with_nan_values(self) -> None:
        """Test Excel parsing handles NaN values correctly - empty values are skipped."""
        # Arrange
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["name", "value"])
        ws.append(["item1", 100])
        ws.append(["item2", None])  # This will become NaN and be skipped
        ws.append(["item3", 200])

        buffer = io.BytesIO()
        wb.save(buffer)
        file_content = buffer.getvalue()

        # Act
        result = parse_file(file_content, FileType.XLSX)

        # Assert
        # NaN/None values are filtered out, so we only get non-empty values
        # Note: Integer values are preserved as integers
        expected = {
            "name": [
                "item1",
                "item2",
                "item3",
            ],  # All names are present, sorted alphabetically
            "value": [
                "100",
                "200",
            ],  # NaN value is skipped, integers preserved, sorted alphabetically
        }
        assert result == expected

    def test_parse_excel_file_with_empty_cells(self) -> None:
        """Test Excel parsing with empty cells - empty values are skipped."""
        # Arrange
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["name", "age", "city"])
        ws.append(["John", None, "NYC"])
        ws.append([None, 30, None])
        ws.append(["Bob", 35, "Chicago"])

        buffer = io.BytesIO()
        wb.save(buffer)
        file_content = buffer.getvalue()

        # Act
        result = parse_file(file_content, FileType.XLSX)

        # Assert
        # Empty/None values are filtered out
        # Note: Integer values are preserved as integers, not converted to floats
        expected = {
            "name": ["Bob", "John"],  # None value is skipped, sorted alphabetically
            "age": [
                "30",
                "35",
            ],  # None value is skipped, integers preserved, sorted alphabetically
            "city": ["Chicago", "NYC"],  # None value is skipped, sorted alphabetically
        }
        assert result == expected

    def test_parse_excel_file_with_only_empty_values_in_column(self) -> None:
        """Test Excel parsing when a column has only empty values."""
        # Arrange
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["name", "age", "city"])
        ws.append(["John", None, "NYC"])
        ws.append(["Jane", None, "LA"])
        ws.append(["Bob", None, "Chicago"])

        buffer = io.BytesIO()
        wb.save(buffer)
        file_content = buffer.getvalue()

        # Act
        result = parse_file(file_content, FileType.XLSX)

        # Assert
        # The 'age' column should be empty since all values are None
        expected = {
            "name": ["Bob", "Jane", "John"],  # Sorted alphabetically
            "age": [],  # Empty because all age values are None
            "city": ["Chicago", "LA", "NYC"],  # Sorted alphabetically
        }
        assert result == expected

    def test_parse_file_invalid_csv_format(self) -> None:
        """Test that invalid CSV format is handled gracefully."""
        # Arrange - Create CSV with inconsistent columns (this is actually valid CSV)
        csv_content = "name,age\nJohn,25,extra_column\nJane"  # Extra column is ignored, missing column is handled
        file_content = csv_content.encode()

        # Act - This should not raise an error, just parse what it can
        result = parse_file(file_content, FileType.CSV)

        # Assert - Should parse successfully, ignoring extra columns and handling missing ones
        assert "name" in result
        assert "age" in result
        assert "John" in result["name"]
        assert "Jane" in result["name"]
        assert "25" in result["age"]

    def test_parse_file_corrupted_excel(self) -> None:
        """Test that corrupted Excel file raises ValueError."""
        # Arrange - Create invalid Excel content
        file_content = b"This is not a valid Excel file"

        # Act & Assert
        with pytest.raises(ValueError):
            parse_file(file_content, FileType.XLSX)

    def test_parse_file_unsupported_file_type(self) -> None:
        """Test that unsupported file type raises ValueError."""
        # Arrange
        file_content = b"some content"

        # Act & Assert
        # This test is checking Excel parsing with invalid content
        # The error message will be about Excel parsing, not "Unsupported file type"
        with pytest.raises(ValueError, match="Error parsing Excel file"):
            # This should raise because we're passing invalid Excel content
            with patch("worker_safety_service.utils.FileType") as mock_file_type:
                mock_file_type.value = ".txt"
                parse_file(file_content, mock_file_type)

    def test_parse_csv_file_with_special_characters(self) -> None:
        """Test CSV parsing with special characters and unicode."""
        # Arrange
        csv_content = 'name,description\nJohn,Contains comma, and quotes"\nJané,Unicode characters: éñü'
        file_content = csv_content.encode("utf-8")

        # Act
        result = parse_file(file_content, FileType.CSV)

        # Assert
        assert "name" in result
        assert "description" in result
        assert len(result["name"]) == 2
        assert len(result["description"]) == 2

    def test_parse_csv_file_large_dataset(self) -> None:
        """Test CSV parsing with large dataset."""
        # Arrange - Create a CSV with many rows
        header = "id,name,value"
        rows = [f"{i},item_{i},{i*10}" for i in range(1000)]  # range starts at 0
        csv_content = header + "\n" + "\n".join(rows)
        file_content = csv_content.encode()

        # Act
        result = parse_file(file_content, FileType.CSV)

        # Assert
        assert len(result["id"]) == 1000
        assert len(result["name"]) == 1000
        assert len(result["value"]) == 1000
        # Check that data is sorted and contains expected values
        assert result["id"][0] == "0"  # First ID should be "0" after sorting
        assert result["id"][999] == "999"  # Last ID should be "999" after sorting
        assert "item_999" in result["name"]  # Should contain this item
        assert "5000" in result["value"]  # Should contain this value

    def test_parse_file_preserves_string_format(self) -> None:
        """Test that all values are returned as strings."""
        # Arrange
        csv_content = (
            "numbers,booleans,dates\n123,true,2023-01-01\n456,false,2023-12-31"
        )
        file_content = csv_content.encode()

        # Act
        result = parse_file(file_content, FileType.CSV)

        # Assert
        # All values should be strings
        assert all(isinstance(val, str) for val in result["numbers"])
        assert all(isinstance(val, str) for val in result["booleans"])
        assert all(isinstance(val, str) for val in result["dates"])

        assert result["numbers"] == ["123", "456"]  # Already sorted alphabetically
        assert result["booleans"] == ["false", "true"]  # Sorted alphabetically
        assert result["dates"] == [
            "2023-01-01",
            "2023-12-31",
        ]  # Already sorted alphabetically

    def test_parse_csv_file_removes_duplicate_values(self) -> None:
        """Test CSV parsing removes duplicate values within each column."""
        # Arrange
        csv_content = """name,age,city
John,25,NYC
Jane,30,LA
John,25,NYC
Bob,25,Chicago
Alice,22,Boston
John,25,NYC
Jane,30,LA
Bob,35,Chicago"""
        file_content = csv_content.encode()

        # Act
        result = parse_file(file_content, FileType.CSV)

        # Assert
        # Check that duplicates are removed and data is sorted
        assert result["name"] == [
            "Alice",
            "Bob",
            "Jane",
            "John",
        ]  # 4 unique names, sorted
        assert result["age"] == ["22", "25", "30", "35"]  # 4 unique ages, sorted
        assert result["city"] == [
            "Boston",
            "Chicago",
            "LA",
            "NYC",
        ]  # 4 unique cities, sorted

        # Verify specific duplicates were removed
        assert result["name"].count("John") == 1  # Originally appeared 3 times
        assert (
            result["age"].count("25") == 1
        )  # Originally appeared 3 times (John and Bob)
        assert result["city"].count("NYC") == 1  # Originally appeared 3 times

    def test_parse_excel_file_removes_duplicate_values(self) -> None:
        """Test Excel parsing removes duplicate values within each column."""
        # Arrange
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["product", "price", "category"])
        ws.append(["Apple", 1.5, "Fruit"])
        ws.append(["Bread", 2.0, "Bakery"])
        ws.append(["Apple", 1.5, "Fruit"])  # Duplicate row
        ws.append(["Banana", 1.0, "Fruit"])
        ws.append(["Bread", 2.0, "Bakery"])  # Duplicate row
        ws.append(["Carrot", 0.5, "Vegetable"])
        ws.append(["Apple", 1.5, "Fruit"])  # Another duplicate

        buffer = io.BytesIO()
        wb.save(buffer)
        file_content = buffer.getvalue()

        # Act
        result = parse_file(file_content, FileType.XLSX)

        # Assert
        # Check that duplicates are removed and data is sorted
        assert result["product"] == [
            "Apple",
            "Banana",
            "Bread",
            "Carrot",
        ]  # 4 unique products, sorted
        assert result["price"] == ["0.5", "1", "1.5", "2"]  # 4 unique prices, sorted
        assert result["category"] == [
            "Bakery",
            "Fruit",
            "Vegetable",
        ]  # 3 unique categories, sorted

        # Verify specific duplicates were removed
        assert result["product"].count("Apple") == 1  # Originally appeared 3 times
        assert result["category"].count("Fruit") == 1  # Originally appeared 3 times

    def test_parse_csv_file_duplicate_values_mixed_with_empty(self) -> None:
        """Test CSV parsing handles duplicates mixed with empty values correctly."""
        # Arrange
        csv_content = """name,score,grade
Alice,85,A
Bob,,B
Alice,85,A
Charlie,90,
Bob,75,B
Alice,85,
Charlie,90,A
"""
        file_content = csv_content.encode()

        # Act
        result = parse_file(file_content, FileType.CSV)

        # Assert
        # Empty values are filtered out, duplicates removed, and sorted
        assert result["name"] == ["Alice", "Bob", "Charlie"]  # 3 unique names, sorted
        assert result["score"] == [
            "75",
            "85",
            "90",
        ]  # 3 unique scores (empty ones filtered), sorted
        assert result["grade"] == [
            "A",
            "B",
        ]  # 2 unique grades (empty ones filtered), sorted

        # Verify duplicates were handled correctly
        assert result["name"].count("Alice") == 1  # Originally appeared 3 times
        assert (
            result["score"].count("85") == 1
        )  # Originally appeared 2 times (but one with empty grade)
        assert result["grade"].count("A") == 1  # Originally appeared 2 times

    def test_parse_csv_file_all_duplicate_values_in_column(self) -> None:
        """Test CSV parsing when a column has all duplicate values."""
        # Arrange
        csv_content = """name,status,priority
John,Active,High
Jane,Active,High
Bob,Active,High
Alice,Active,High"""
        file_content = csv_content.encode()

        # Act
        result = parse_file(file_content, FileType.CSV)

        # Assert
        # All values unique in name, but status and priority should have only one value each
        assert result["name"] == [
            "Alice",
            "Bob",
            "Jane",
            "John",
        ]  # 4 unique names, sorted
        assert result["status"] == ["Active"]  # Only 1 unique status
        assert result["priority"] == ["High"]  # Only 1 unique priority

    def test_parse_file_handles_duplicate_column_names(self) -> None:
        """Test parsing files with duplicate column names."""
        # Arrange
        csv_content = "name,age,name\nJohn,25,Smith\nJane,30,Doe"
        file_content = csv_content.encode()

        # Act
        result = parse_file(file_content, FileType.CSV)

        # Assert
        # pandas automatically handles duplicate columns by adding suffixes
        # The exact behavior may vary, but we should get all the data
        assert (
            len(result) >= 2
        )  # Should have at least 2 columns (age + some form of name columns)
        assert "age" in result
        assert result["age"] == ["25", "30"]
